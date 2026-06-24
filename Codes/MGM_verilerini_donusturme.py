import openpyxl

def read_meteorology_data_to_excel(excel_path, output_path):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Yeni bir workbook oluştur
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Veriler"

    # Başlıkları sadece ilk istasyon için yaz
    header_written = False

    # İstasyon isimlerini izlemek için bir set oluştur
    processed_stations = set()

    # Sayfa adlarını dolaş
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]  
        print(f"\nSayfa: {sheet_name}")
        
        rows = list(sheet.iter_rows(values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]

            # "İstasyon Adı/No:" içeren satır
            if any(cell and "İstasyon Adı/No:" in str(cell) for cell in row):
                station_cell = next(cell for cell in row if cell and "İstasyon Adı/No:" in str(cell))
                station_name = str(station_cell).split("İstasyon Adı/No:")[1].split('/')[0].strip()

                # Eğer istasyon zaten işlendi ise geç
                if station_name in processed_stations:
                    i += 1
                    continue

                processed_stations.add(station_name)
                print(f"\nİSTASYON: {station_name}")
                
                # Başlıkları bul
                header = []
                for j in range(i + 1, i + 6):  # En fazla 5 satır arar
                    if j < len(rows):
                        possible_header = rows[j]
                        if any(possible_header) and any("Yıl" in str(cell) or "Ay" in str(cell) for cell in possible_header if cell):
                            header = [str(cell).strip() if cell is not None else "" for cell in possible_header]
                            header_row_index = j
                            break

                if not header:
                    print("Tablo başlığı bulunamadı!")
                    i += 1
                    continue

                # "Aylık Ortalama Nispi Nem (%)" sütununu atla
                header = [h for h in header if h != "Aylık Ortalama Nispi Nem (%)"]

                # Verileri oku ve Excel'e yaz (ilk satırı atla)
                for k in range(header_row_index + 2, len(rows)):
                    data_row = rows[k]

                    if all(cell is None for cell in data_row):
                        break  # Tamamen boşsa dur

                    # Satırda yeterli veri olup olmadığına bak
                    cleaned_row = [station_name]  # İstasyon adını başta ekle

                    # Diğer sütunlara veri ekle (1-12 sütunlarını al, sadece aylar)
                    for idx, cell in enumerate(data_row[2:14]):  # Sadece 1-12 sütunlarını al (Aylar)
                        if isinstance(cell, float):
                            cleaned_row.append(f"{cell:.1f}")
                        elif cell is None:
                            cleaned_row.append("")
                        else:
                            cleaned_row.append(str(cell).strip())
                    
                    output_sheet.append(cleaned_row)

                i = k  # Devam etmek için indeks güncelle
            else:
                i += 1

    workbook.close()

    # Yeni Excel dosyasını kaydet
    output_workbook.save(output_path)
    print(f"\nVeriler başarıyla {output_path} dosyasına kaydedildi.")

# Kullanım
excel_file = "Sicaklik.xlsx"
output_file = "Sicaklik_Output_Data.xlsx"
read_meteorology_data_to_excel(excel_file, output_file)
